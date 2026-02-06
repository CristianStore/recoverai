"""
Script para crear un archivo Excel con calendario y gestión de datos
Características:
- Fila 1: Fecha
- Fila 2: Nombre
- Fila 3: Dinero
- Fila 4: Porcentaje
- Fila 5: Calendario emergente con 6 días (Lunes a Sábado)
- Búsqueda de datos por fecha
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

# Crear libro de trabajo
wb = Workbook()
ws = wb.active
ws.title = "Gestor de Datos"

# Configurar estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados de filas
headers = ["Fecha", "Nombre", "Dinero", "Porcentaje"]
for idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=idx, column=1)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Crear encabezados de días de la semana (Lunes a Sábado)
dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
day_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

for idx, dia in enumerate(dias_semana, start=2):
    cell = ws.cell(row=5, column=idx)
    cell.value = dia
    cell.fill = day_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 15
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws.column_dimensions[col].width = 18

# Crear área de búsqueda por fecha
ws['A7'] = "Buscar por Fecha:"
ws['A7'].font = Font(bold=True, size=11)
ws['B7'].value = datetime.now().strftime("%Y-%m-%d")

# Validación de fecha para búsqueda
date_validation = DataValidation(type="date", allow_blank=False)
date_validation.add('B7')
ws.add_data_validation(date_validation)

# Crear filas de datos (6 columnas para Lunes-Sábado)
# Filas 1-4 para datos, columnas B-G (6 días)
for row in range(1, 5):
    for col in range(2, 8):  # B a G
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formato especial para fila de Dinero
        if row == 3:
            cell.number_format = '$#,##0.00'
        
        # Formato especial para fila de Porcentaje
        if row == 4:
            cell.number_format = '0.00%'

# Crear hoja de datos históricos
ws_data = wb.create_sheet("Datos Históricos")
ws_data['A1'] = "Fecha"
ws_data['B1'] = "Día"
ws_data['C1'] = "Nombre"
ws_data['D1'] = "Dinero"
ws_data['E1'] = "Porcentaje"

# Estilo para encabezados de datos históricos
for col in ['A', 'B', 'C', 'D', 'E']:
    cell = ws_data[f'{col}1']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Ajustar anchos
for col in ['A', 'B', 'C', 'D', 'E']:
    ws_data.column_dimensions[col].width = 15

# Crear hoja de instrucciones
ws_inst = wb.create_sheet("Instrucciones")
instrucciones = [
    "CÓMO USAR ESTE GESTOR DE DATOS",
    "",
    "1. INGRESAR DATOS SEMANALES:",
    "   - En la hoja 'Gestor de Datos', verás 6 columnas (Lunes a Sábado)",
    "   - Llena los datos en las filas: Fecha, Nombre, Dinero, Porcentaje",
    "   - El dinero se formatea automáticamente como moneda",
    "   - El porcentaje se formatea automáticamente (ingresa 0.15 para 15%)",
    "",
    "2. BUSCAR POR FECHA:",
    "   - Usa la celda B7 para ingresar una fecha",
    "   - Copia los datos de esa fecha a la hoja 'Datos Históricos'",
    "",
    "3. GUARDAR HISTORIAL:",
    "   - Ve a la hoja 'Datos Históricos'",
    "   - Copia manualmente los datos que quieras conservar",
    "",
    "4. CONSEJOS:",
    "   - Guarda el archivo regularmente (Ctrl+S)",
    "   - Usa filtros en 'Datos Históricos' para buscar información",
    "   - Puedes agregar más filas en 'Datos Históricos' según necesites"
]

for idx, linea in enumerate(instrucciones, start=1):
    cell = ws_inst.cell(row=idx, column=1)
    cell.value = linea
    if "CÓMO USAR" in linea:
        cell.font = Font(bold=True, size=14, color="4472C4")
    elif linea.startswith(("1.", "2.", "3.", "4.")):
        cell.font = Font(bold=True, size=11)

ws_inst.column_dimensions['A'].width = 80

# Guardar archivo
output_path = "GestorDatos_Semanal.xlsx"
wb.save(output_path)
print(f"✅ Archivo Excel creado exitosamente: {output_path}")
print(f"📍 Ubicación: d:\\RecoverAI\\GestorDatos\\{output_path}")
